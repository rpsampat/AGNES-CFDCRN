import numpy
import pickle
from pyevtk.hl import *
from openpyxl import Workbook
import numpy as np


def save_data(save_dict, name, path):
    """
    Saves data in save_dict such that data under each key occupies a separate column in a .xls file.
    :param save_dict: dictionary {1: list1, 2: list2...}
    :param name: file name without extension
    :param path: path to be saved in
    :return:
    """
    loc = path
    dest_filename = loc + '/%s.xlsx' % (name)
    wb = Workbook()

    # General data
    ws = wb.active
    ws.title = 'General'
    row = 0
    cols = len(save_dict.keys())
    for i in range(len(save_dict[1])):
        row += 1
        for j in range(cols):
            ws.cell(column=j + 1, row=row, value=save_dict[j + 1][i])

    wb.save(filename=dest_filename)


def thermal_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'N + NO <=> N2 + O'
    reac2 = 'N + O2 <=> NO + O'
    reac3 = 'N + OH <=> H + NO'
    for i in range(number):
        k1 = 0
        k2 = 0
        k = 0
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2 or str == reac3:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def n2o_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'N2O (+M) <=> N2 + O (+M)'
    reac2 = 'N2O + O <=> 2 NO'

    for i in range(number):

        str = g.reaction_equation(i)
        if str == reac1 or str == reac2:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def nnh_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'NNH <=> H + N2'
    reac2 = 'NNH + O <=> NH + NO'

    for i in range(number):

        str = g.reaction_equation(i)
        if str == reac1 or str == reac2:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def save(reactors, header, data, mass_imb, outlet_react, inlet_react, tres, path, identity, data_num, symm_axis):
    loc = path
    dest_filename = loc + '/data%i.xlsx' % identity
    wb = Workbook()

    # General data
    ws = wb.active
    ws.title = 'General'
    NO_ind = 0
    NO2_ind = 0
    O2_ind = 0
    H2O_ind = 0

    col = 1
    row = 1
    ws.cell(column=col, row=row, value='Reactor name')
    col += 1
    ws.cell(column=col, row=row, value='Temperature [K]')
    col += 1
    ws.cell(column=col, row=row, value='Pressure [Pa]')
    col += 1
    ws.cell(column=col, row=row, value='Residence time [ms]')
    col += 1
    ws.cell(column=col, row=row, value='Mass [kg]')
    col += 1
    ws.cell(column=col, row=row, value='Volume [m^3]')
    col += 1
    ws.cell(column=col, row=row, value='Density [kg/m^3]')
    col += 1
    ws.cell(column=col, row=row, value='Mean Molecular Weight []')
    col += 1
    ws.cell(column=col, row=row, value='Specific Heat Capacity [J/kgK]')
    col += 1
    ws.cell(column=col, row=row, value='Thermal NOx')
    col += 1
    ws.cell(column=col, row=row, value='N2O pathway')
    col += 1
    ws.cell(column=col, row=row, value='NNH pathway')

    for sp in reactors[0].thermo.species_names:
        col += 1
        ws.cell(column=col, row=row, value=sp)
        if sp == 'NO':
            NO_ind = col
        if sp == 'NO2':
            NO2_ind = col
        if sp == 'O2':
            O2_ind = col
        if sp == 'H2O':
            H2O_ind = col
    for sp in reactors[0].thermo.species_names:
        col += 1
        ws.cell(column=col, row=row, value='Mass fraction of ' + sp)

    Pressure = []
    Volume = []
    Temperature = []
    Mass = []
    O2 = []
    NO = []
    NO2 = []
    H2O = []

    row += 1
    for r_len in range(len(reactors)):
        r = reactors[r_len]
        col = 1
        ws.cell(column=col, row=row, value=r.name)
        col += 1
        ws.cell(column=col, row=row, value=r.T)
        #print "Temp = ", r.T
        Temperature.append(r.T)
        col += 1
        ws.cell(column=col, row=row, value=r.thermo.P)
        Pressure.append(r.thermo.P)
        col += 1
        try:
            ws.cell(column=col, row=row, value=tres[reactors.index(r)])
        except:
            ind = np.where(reactors == r)
            ws.cell(column=col, row=row, value=tres[ind[0][0]])
        col += 1
        ws.cell(column=col, row=row, value=r.mass)
        Mass.append(r.mass)
        col += 1
        ws.cell(column=col, row=row, value=r.volume)
        Volume.append(r.volume)
        col += 1
        ws.cell(column=col, row=row, value=r.density)
        col += 1
        ws.cell(column=col, row=row, value=r.thermo.mean_molecular_weight)
        col += 1
        ws.cell(column=col, row=row, value=r.thermo.cp)
        col += 1
        ws.cell(column=col, row=row, value=thermal_pathway(r))
        col += 1
        ws.cell(column=col, row=row, value=n2o_pathway(r))
        col += 1
        ws.cell(column=col, row=row, value=nnh_pathway(r))
        for x in r.thermo.X:
            col += 1
            ws.cell(column=col, row=row, value=x)
            if col == NO_ind:
                NO.append(x)
            if col == NO2_ind:
                NO2.append(x)
            if col == O2_ind:
                O2.append(x)
            if col == H2O_ind:
                H2O.append(x)
        for y in r.thermo.Y:
            col += 1
            ws.cell(column=col, row=row, value=y)
        row += 1

    wb.save(filename=dest_filename)
    x = header.index('X-Coordinate') + 1
    y = header.index('Y-Coordinate') + 1
    z = header.index('Z-Coordinate') + 1

    X = numpy.array(data[data_num][x], dtype=numpy.float64)
    Y = numpy.array(data[data_num][y], dtype=numpy.float64)
    Z = numpy.array(data[data_num][z], dtype=numpy.float64)

    dat = {}
    for sp in reactors[0].thermo.species_names:
        temp = []
        index = reactors[0].thermo.species_names.index(sp)
        for r in reactors:
            temp.append(r.thermo.Y[index])
        dat[sp] = numpy.array(temp, dtype=numpy.float64)

    temp1 = []
    temp2 = []
    temp3 = [0] * len(reactors)
    temp4 = [0] * len(reactors)
    temp5 = []
    temp6 = []
    temp7 = []
    temp8 = []

    for r in reactors:
        temp1.append(r.T)
        temp2.append(r.thermo.P)
        temp5.append(r.density)
        temp6.append(r.volume)
        temp7.append(r.thermo.mean_molecular_weight)
        temp8.append(r.thermo.cp)
    for i in inlet_react:
        temp3[i] = 1
    for o in outlet_react:
        temp4[o] = 1

    try:
        temp9 = numpy.array(data[data_num][header.index('Axial Velocity') + 1]) * \
                numpy.array(data[data_num][header.index('Density') + 1]) / numpy.array(temp5)
        temp10 = numpy.array(data[data_num][header.index('Radial Velocity') + 1]) * \
                 numpy.array(data[data_num][header.index('Density') + 1]) / numpy.array(temp5)
    except ValueError:
        print ''

    with open(path + '/outlet.pkl', 'wb') as file:
        pickle.dump(outlet_react, file, pickle.HIGHEST_PROTOCOL)
    dat['Temperature'] = numpy.array(temp1, dtype=numpy.float64)
    dat['Pressure'] = numpy.array(temp2, dtype=numpy.float64)
    try:
        dat['MassImbalance'] = numpy.array(mass_imb, dtype=numpy.float64)
    except:
        print ''
    dat['Inlet'] = numpy.array(temp3, dtype=numpy.float64)
    dat['Outlet'] = numpy.array(temp4, dtype=numpy.float64)
    dat['Density'] = numpy.array(temp5, dtype=numpy.float64)
    dat['Volume'] = numpy.array(temp6, dtype=numpy.float64)
    dat['Molecular Weight'] = numpy.array(temp7, dtype=numpy.float64)
    dat['Specific Heat Capacity'] = numpy.array(temp8, dtype=numpy.float64)
    dat['Residence Time'] = numpy.array(tres, dtype=numpy.float64)
    try:
        dat['Turbulent Viscosity'] = numpy.array(data[data_num][header.index('Turbulent Viscosity') + 1],
                                                 dtype=numpy.float64)
    except:
        print ''

    try:
        dat['Axial Velocity'] = numpy.array(temp9, dtype=numpy.float64)
        dat['Radial Velocity'] = numpy.array(temp10, dtype=numpy.float64)
    except:
        print ''
    pointsToVTK(path + "/para_CRN%i" % identity, X, Y, Z, data=dat)

    NOx = []
    R = 8.314
    perc_corr = 0.15  # O2 correction

    for conc in range(len(NO)):
        P = Pressure[conc]
        T = Temperature[conc]
        V = Volume[conc]
        N_wet = P * V / (R * T)
        N_dry = N_wet * (1 - H2O[conc])
        NO_dry = NO[conc] * N_wet / N_dry
        # O2 correction
        N_O2 = O2[conc] * N_wet
        O2_corr = (perc_corr / (1 - perc_corr)) * (N_dry - N_O2)
        N_corr = N_dry - N_O2 + O2_corr
        NO_corr = NO_dry * N_dry / N_corr
        NOx.append(NO_corr * 1e6)

    z2plot = {}
    NOxplot = {}
    mass_react = {}

    x = header.index(symm_axis) + 1
    try:
        for ind in range(len(data[data_num][x])):
            xcoord = data[data_num][x][ind]
            if xcoord in NOxplot:
                x_id = xcoord
                NOxplot[x_id] = (NOxplot[x_id] * mass_react[x_id] + NOx[ind] * Mass[ind]) / (
                            mass_react[x_id] + Mass[ind])
                mass_react[x_id] = mass_react[x_id] + Mass[ind]
            else:
                NOxplot[xcoord] = NOx[ind]
                mass_react[xcoord] = Mass[ind]
    except:
        a=1
    # all coordinates currently in metre
    dict_sort = sorted(NOxplot.keys())
    limit = 0.010  # 10mm
    startpt = 0
    mass_react2 = {}

    for pt in dict_sort:
        if pt > 0:
            if pt < limit:
                try:
                    z2plot[startpt] = (z2plot[startpt] * mass_react2[startpt] + NOxplot[pt] * mass_react[pt]) / \
                                      (mass_react2[startpt] + mass_react[pt])
                    mass_react2[startpt] = mass_react2[startpt] + Mass[pt]
                except:
                    z2plot[startpt] = NOxplot[pt]
                    mass_react2[startpt] = mass_react[pt]
            else:
                while pt > limit:
                    startpt += limit
                    limit += startpt
                z2plot[startpt] = NOxplot[pt]
                mass_react2[startpt] = mass_react[pt]

    # write NOx data to file
    dest_filename = loc + '/Emissions%i.xlsx' % identity
    wb = Workbook()

    ws = wb.active
    ws.title = 'NOx'
    row = 1
    ws.cell(column=1, row=row, value=symm_axis[0] + ' (mm)')
    ws.cell(column=2, row=row, value='NOx ppm')
    row += 1

    for j in z2plot.keys():
        if j > 0:
            ws.cell(column=1, row=row, value=j * 1000)
            ws.cell(column=2, row=row, value=z2plot[j])
            row += 1
    wb.save(filename=dest_filename)
