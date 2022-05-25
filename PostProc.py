from openpyxl import load_workbook
import math
import pickle
import cantera as ct
import numpy
import Read_Plot_all as rpa
import matplotlib.pyplot as plt
from matplotlib.gridspec import GridSpec
from pyevtk.hl import*


def radical_consump(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'O + CH <=> H + CO'
    reac2 = 'O + CH2 <=> H + HCO'
    reac3 = 'O + CH3 <=> H + CH2O'
    reac4 = 'O + CH3 <=> H + H2 + CO'
    for i in range(number):
        k1 = 0
        k2 = 0
        k = 0
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2 or str == reac3 or str == reac4:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('O')
                ind = g.kinetics_species_index('O')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('O')
                    ind = g.kinetics_species_index('O')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def prompt_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'CH + N2 <=> HCN + N'
    reac2 = 'CH + N2 (+M) <=> HCNN (+M)'
    reac3 = 'CH2 + N2 <=> HCN + NH'
    reac4 = 'H2CN + N <=> N2 + CH2'
    reac5 = 'HCNN + H <=> CH2 + N2'
    reac6 = 'CH2(S) + N2 <=> NH + HCN'

    for i in range(number):
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2 or str == reac3 or str == reac4 or str == reac5 or str == reac6:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('N2')
                ind = g.kinetics_species_index('N2')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('N2')
                    ind = g.kinetics_species_index('N2')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def reburn_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'CH + NO <=> HCN + O'
    reac2 = 'CH + NO <=> H + NCO'
    reac3 = 'CH + NO <=> N + HCO'
    reac4 = 'CH3 + NO <=> HCN + H2O'
    reac5 = 'CH3 + NO <=> H2CN + OH'
    reac6 = 'CH2 + NO <=> OH + HCN'
    reac7 = 'CH2 + NO <=> H + HNCO'
    reac8 = 'CH2 + NO <=> H + HCNO'
    reac9 = 'CH2(S) + NO <=> H + HNCO'
    reac10 = 'CH2(S) + NO <=> OH + HCN'
    reac11 = 'CH2(S) + NO <=> H + HCNO'
    reac12 = 'HCCO + NO <=> HCNO + CO'
    reac13 = 'C + NO <=> CN + O'
    reac14 = 'C + NO <=> CO + N'

    for i in range(number):
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2 or str == reac3 or str == reac4 or str == reac5 or str == reac6 or str == reac7\
                or str == reac8 or str == reac9 or str == reac10 or str == reac11 or str == reac12 or str == reac13 \
                or str == reac14:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def thermal_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'N + NO <=> N2 + O'
    for i in range(number):
        str = g.reaction_equation(i)
        if str == reac1:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def n2o_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'N2O (+M) <=> N2 + O (+M)'
    reac2 = 'N2O + O <=> 2 NO'
    for i in range(number):
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def nnh_pathway(PSR):
    # gas in PSR
    g = PSR.thermo
    # number of reactions
    number = g.n_reactions
    k_tot = 0
    reac1 = 'NNH <=> H + N2'
    reac2 = 'NNH + O <=> NH + NO'

    for i in range(number):
        str = g.reaction_equation(i)
        if str == reac1 or str == reac2:
            react = g.reactants(i)
            react = react.split(' ')
            prod = g.products(i)
            prod = prod.split(' ')
            try:
                # species in product
                check = prod.index('NO')
                ind = g.kinetics_species_index('NO')
                coeff = g.product_stoich_coeff(ind, i)
                k = g.net_rates_of_progress[i] * coeff
                k_tot += k
            except:
                try:
                    # species in reactant
                    check = react.index('NO')
                    ind = g.kinetics_species_index('NO')
                    coeff = g.reactant_stoich_coeff(ind, i)
                    k = -g.net_rates_of_progress[i] * coeff
                    k_tot += k
                except:
                    continue

    return k_tot


def Post(path, data_num, symm_axis):
    print 'Start'
    with open(path+"/graph2plot.pkl", 'rb') as f:
        graph = pickle.load(f)
    with open("header.pkl", 'rb') as f:
        header = pickle.load(f)
    with open("data.pkl", 'rb') as f:
        data = pickle.load(f)
    with open('cellaxsurf.pkl', 'rb') as f:
        cell_ax_surf = pickle.load(f)

    # theta_index = header.index('Abs. Angular Coordinate')+1
    # theta_max = max(data[data_num][theta_index])
    # theta_min = min(data[data_num][theta_index])
    # print theta_index
    # print theta_max
    # print theta_min
    # define normal of plane ax+by+cz=0
    a = 0
    b = 0
    c = 1

    cells2plot = {}
    x_ind = header.index('X-Coordinate')+1
    y_ind = header.index('Y-Coordinate')+1
    z_ind = header.index('Z-Coordinate')+1
    Queue = []
    diction = {'X': [], 'Y': [], 'Z': [], 'Reactor': []}
    print len(data[data_num][x_ind])
    for i in range(len(data[data_num][x_ind])):
        x = data[data_num][x_ind][i]
        y = data[data_num][y_ind][i]
        z = data[data_num][z_ind][i]
        resid = a*x+b*y+c*z
        if abs(resid) < 1e-4:
            if symm_axis == 'X-Coordinate':
                cells2plot[i+1] = [math.sqrt(z**2+y**2), x]
            elif symm_axis == 'Y-Coordinate':
                cells2plot[i+1] = [math.sqrt(z**2+x**2), y]
            elif symm_axis == 'Z-Coordinate':
                cells2plot[i+1] = [math.sqrt(x**2+y**2), z]
            Queue.append(i+1)

    wb = load_workbook(path+'/data4.xlsx')
    ws = wb['General']
    print len(cells2plot)
    print "Searching cell"
    titlelist = {}
    species = ['H2', 'H', 'O', 'O2', 'OH', 'H2O', 'HO2', 'H2O2', 'C', 'CH', 'CH2', 'CH2(S)', 'CH3', 'CH4', 'CO', 'CO2',
               'HCO', 'CH2O', 'CH2OH', 'CH3O', 'CH3OH', 'C2H', 'C2H2', 'C2H3', 'C2H4', 'C2H5', 'C2H6', 'HCCO', 'CH2CO',
               'HCCOH', 'N', 'NH', 'NH2', 'NH3', 'NNH', 'NO', 'NO2', 'N2O', 'HNO', 'CN', 'HCN', 'H2CN', 'HCNN', 'HCNO',
               'HOCN', 'HNCO', 'NCO', 'N2', 'AR', 'C3H7', 'C3H8', 'CH2CHO', 'CH3CHO']
    y_list = {}
    vol_id = []
    temp_id = []
    p_id = []
    for number in range(ws.max_column-1):
        val = ws.cell(column=number+2, row=1)
        title = str(val.value)
        title = title.replace(" ", "")
        if title in species:
            y_list[title] = number+2
        if title == "Volume[m^3]":
            vol_id.append(number+2)
        if title == "Temperature[K]":
            temp_id.append(number + 2)
        if title == "Pressure[Pa]":
            p_id.append(number+2)
        if title.find("Mass fraction of") == -1 and title.find("pathway") == -1:
            titlelist[title] = number+2
            diction[title] = []
    diction['MMW'] = []
    diction['Thermal NOx'] = []
    diction['Prompt NOx'] = []
    diction['Reburn NOx'] = []
    diction['NOx Destr Rate'] = []
    diction['O Alt consump'] = []
    diction['NOx Net Rate(bin)'] = []
    diction['NOx Net Rate'] = []
    diction['Velocity dir'] = []

    diction['Axial Velocity[m/s]'] = []
    diction['Cell Volume[m^3]'] = []

    cell2react = {}
    gas = ct.Solution('gri30.cti')
    psr = ct.IdealGasReactor(gas)
    print "Creating new graph"
    thermalnoxrate = []
    promptnoxrate = []
    reburnnoxrate = []
    destnox = []
    consumption = []
    netnox = []
    netnoxbin = []
    for reactor in graph:
        cells = graph[reactor]['cells']
        for c1 in cells:
            cell2react[c1] = reactor
        y_react = {}
        for y in y_list:
            y_react[y] = float(ws.cell(column=y_list[y], row=reactor+1).value)
        t = float(ws.cell(column=temp_id[0], row=reactor+1).value)
        pr = float(ws.cell(column=p_id[0], row=reactor + 1).value)
        vol = float(ws.cell(column=vol_id[0], row=reactor + 1).value)
        gas.TPX = t, pr, y_react
        psr.volume = vol
        psr.insert(gas)
        noxdestr = gas.destruction_rates[species.index('NO')]
        noxnet = gas.net_production_rates[species.index('NO')]
        destnox.append(noxdestr + 1e-20)

        # net nox rate
        netrate = 0
        netnox.append(noxnet)
        if noxnet > 0:
            netrate = 1
        elif noxnet < 0:
            netrate = -1
        netnoxbin.append(netrate)

        # thermal nox
        thermalpath = thermal_pathway(psr)
        tnox = 1e-15
        if thermalpath > 0:
            tnox = thermalpath
        thermalnoxrate.append(tnox)

        # promptnox
        ppath = prompt_pathway(psr)
        phcn = 1e-20
        if ppath < 0:
            phcn = -ppath
        promptnoxrate.append(phcn)

        # nox reburn
        reburnpath = reburn_pathway(psr)
        rbnr = 1e-20
        if reburnpath < 0:
            rbnr = -reburnpath
        reburnnoxrate.append(rbnr)

        # O radical consumption
        radicconsump = radical_consump(psr)
        rcr = 1e-20
        if radicconsump < 0:
            rcr = -radicconsump
        consumption.append(rcr)

    print "Traversing Queue "
    for i in Queue:
        reactor = cell2react[i]
        for quant in titlelist:
            r = titlelist[quant]
            val = ws.cell(row=reactor+1, column=r)
            title = quant
            diction[title].append(val.value)
        diction['O Alt consump'].append(consumption[reactor - 1])
        diction['NOx Destr Rate'].append(destnox[reactor - 1])
        diction['NOx Net Rate'].append(netnox[reactor - 1])
        diction['NOx Net Rate(bin)'].append(netnoxbin[reactor - 1])
        diction['Thermal NOx'].append(thermalnoxrate[reactor - 1])
        diction['Prompt NOx'].append(promptnoxrate[reactor - 1])
        diction['Reburn NOx'].append(reburnnoxrate[reactor - 1])
        diction['X'].append(cells2plot[i][bool(symm_axis == 'X-Coordinate')]*bool(a != 1))
        diction['Y'].append(cells2plot[i][bool(symm_axis == 'Y-Coordinate')]*bool(b != 1))
        diction['Z'].append(cells2plot[i][bool(symm_axis == 'Z-Coordinate')]*bool(c != 1))
        diction['Reactor'].append(reactor)
        index = header.index('Mean Molecular Weight') + 1
        mw = data[data_num][index][i - 1]
        diction['MMW'].append(mw)

        velocity_dict = {'X-Coordinate': ['X Velocity', 'Mesh X-Velocity'],
                         'Y-Coordinate': ['Y Velocity', 'Mesh Y-Velocity'],
                         'Z-Coordinate': ['Z Velocity', 'Mesh Z-Velocity']}
        for coordinate in velocity_dict[symm_axis]:
            if coordinate in header:
                index = header.index('Mesh X-Velocity') + 1
                vel_dir = data[data_num][index][i - 1]
        try:
            diction["Velocity dir"].append(vel_dir/abs(vel_dir))
        except ZeroDivisionError:
            diction['Velocity dir'].append(0)

        # TODO: find a better way to obtain axial velocity
        index = header.index('Velocity Magnitude') + 1
        vel_mag = data[data_num][index][i - 1]
        index = header.index('Radial Velocity') + 1
        vel_rad = data[data_num][index][i - 1]
        diction['Axial Velocity[m/s]'].append(math.sqrt(vel_mag**2-vel_rad**2))

        index = header.index('Cell Volume') + 1
        cell_vol = data[data_num][index][i - 1]
        diction['Cell Volume[m^3]'].append(cell_vol)

    matrix = []
    matrix_header = []

    for cell in range(len(diction['X'])):
        sub_matrix = []
        for head in diction.keys():
            if cell == 0:
                matrix_header.append(head)
            sub_matrix.append(diction[head][cell])
        matrix.append(sub_matrix)

    matrix = sorted(matrix, key=lambda a_entry: float(str(a_entry[matrix_header.index('Y')])[:6]))
    matrix = sorted(matrix, key=lambda a_entry: float(str(a_entry[matrix_header.index('X')])[:6]))

    diction.clear()

    for head in matrix_header:
        diction[head] = []

    for row in matrix:
        for column in range(len(row)):
            diction[matrix_header[column]].append(row[column])

    diction['Cell Axial Surface[m^2]'] = []

    for coor in diction['Y']:
        diction['Cell Axial Surface[m^2]'].append(cell_ax_surf[float(str(coor)[:6])])

    Cluster_min = {}
    Cluster_max = {}
    diction['Cluster'] = []

    for coor in range(len(diction['X'])):
        if diction['Reactor'][coor] not in Cluster_max.keys():
            Cluster_min[diction['Reactor'][coor]] = diction['X'][coor]
        Cluster_max[diction['Reactor'][coor]] = diction['X'][coor]

    for coor in range(len(diction['X'])):
        diction['Cluster'].append(Cluster_max[diction['Reactor'][coor]]-Cluster_min[diction['Reactor'][coor]])

    print "Writing file"
   
    print len(diction['X'])
    print len(diction['Y'])
    print len(diction['Z'])
    print len(diction['Temperature[K]'])
    for key in diction.keys():
        diction[key] = numpy.array(diction[key], dtype=numpy.float64)

    pointsToVTK(path+"/PostProc", diction['X'], diction['Y'], diction['Z'], data=diction)

    return diction


def axial_plot(path, variable, dictions, axial_locations, symm_axis, result_legend, integrals):

    all_data, all_info, all_titles, category = rpa.read_pmd_stat()
    if variable == 'Axial Velocity[m/s]':
        all_data, all_info, all_titles, category = rpa.read_tud_ldv()
    result_legend.append('EXP')

    colors = ['C0', 'C1', 'C4', 'C5', 'C6', 'C9', 'C7', 'C8']

    for axial_loc in axial_locations:
        fig, axes = plt.subplots(2, 2, gridspec_kw={'width_ratios': [11, 1],
                                                    'height_ratios': [len(dictions)+2, 11-len(dictions)]})
        axes[1, 1].set_xticks([])
        axes[1, 1].yaxis.tick_right()
        gs = GridSpec(2, 2, figure=fig, width_ratios=[11, 1], height_ratios=[len(dictions)+2, 9-len(dictions)],
                      wspace=0.2)
        axes[0, 0].remove()
        axes[1, 0].remove()
        axes[0, 1].remove()
        ax1 = fig.add_subplot(gs[0:, 0])
        count = -1

        for diction in dictions:
            count += 1
            saved_axial_base = []
            saved_axial_top = []
            axial_top = None
            axial_base = None

            for j in range(len(diction[symm_axis[0]])-1):
                if diction[symm_axis[0]][j] <= axial_loc < diction[symm_axis[0]][j+1]:
                    axial_base = diction[symm_axis[0]][j]
                    axial_top = diction[symm_axis[0]][j+1]
                    axial_perc = (axial_loc-diction[symm_axis[0]][j])/(diction[symm_axis[0]][j+1]
                                                                       - diction[symm_axis[0]][j])
            for j in range(len(diction[symm_axis[0]])-1):
                if str(diction[symm_axis[0]][j])[:6] == str(axial_base)[:6]:
                    saved_axial_base.append(j)
                elif str(diction[symm_axis[0]][j])[:6] == str(axial_top)[:6]:
                    saved_axial_top.append(j)

            x_axis1 = []
            x_axis2 = []
            x_axis = []

            y_axis1 = []
            y_axis2 = []
            y_axis = []

            axis_options = ['X', 'Y', 'Z']
            for axis in axis_options:
                if axis != symm_axis[0] and max(diction[axis]) != 0:
                    ax = axis

            for cell in saved_axial_base:
                x_axis1.append(diction[ax][cell])
                y_axis1.append(diction[variable][cell])
            for cell in saved_axial_top:
                x_axis2.append(diction[ax][cell])
                y_axis2.append(diction[variable][cell])
            for coor in range(len(x_axis1)):
                x_axis.append(((x_axis2[coor] - x_axis1[coor]) * axial_perc + x_axis1[coor])/0.0072)
                y_axis.append((y_axis2[coor] - y_axis1[coor]) * axial_perc + y_axis1[coor])

            title_dict = {0.0072: 'D01', 0.0144: 'D02', 0.0216: 'D03', 0.108: 'D15', 0.216: 'D30', 0.324: 'D45',
                          0.432: 'D60', 0.54: 'D75'}

            axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                         'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                         'MassfractionofCH4': 'YCH4', 'Axial Velocity[m/s]': 'U(ms-1)'}

            ax1.plot(x_axis, y_axis, '.', color=colors[count])
            axes[1, 1].plot([0, 1], integrals[2][axial_locations.index(axial_loc) + len(axial_locations) * count]
                         * numpy.ones(2), color=colors[count])
        ax1.set_xlabel('Relative radial location (r/d)')
        if variable != 'Temperature[K]' and variable != 'Axial Velocity[m/s]':
            ax1.set_ylabel(variable[:4] + ' ' + variable[4:12] + ' ' + variable[12:14] + ' ' + variable[14:] + ' ($Y_{' +
                           variable[14:] + '}$)')
            axes[1, 1].set_ylabel('$Y_{' + variable[14:] + '}$' + ' volume integral average')
        else:
            ax1.set_ylabel(variable[:11] + ' ' + variable[11:])
            axes[1, 1].set_ylabel('T[K] volume integral average')
        ax1.set_title(symm_axis[0] + '=' + str(round(axial_loc * 1000., 1)) + 'mm (' + title_dict[axial_loc] + ')')

        if variable == 'MassfractionofNO':
            ax1.set_ylim((0, 0.00005))

        if variable == 'MassfractionofCH4' and axial_loc == 0.54:
            ax1.set_ylim((0, 0.0001))

        r_d = ['r/d', 'r/d']
        if variable == 'Axial Velocity[m/s]':
            r_d = ['R/D', 'r/d']
        rpa.plot_desired([title_dict[axial_loc]], r_d, [axis_dict[variable]], category, all_titles, all_data,
                         all_info, rpa.folder_bib_comp, 'radial', path+'/', result_legend,
                         [integrals[0][axial_locations.index(axial_loc)+len(axial_locations)*count],
                          integrals[1][axial_locations.index(axial_loc)+len(axial_locations)*count]],
                         ax1, axes[1, 1], fig)

        plt.gcf().clear()

    return


def radial_plot(path, variable, dictions, radial_loc, symm_axis, result_legend):

    all_data, all_info, all_titles, category = rpa.read_pmd_stat()
    if variable == 'Axial Velocity[m/s]':
        all_data, all_info, all_titles, category = rpa.read_tud_ldv()
    result_legend.append('EXP')
    colors = ['C0', 'C1', 'C4', 'C5', 'C6', 'C9']
    count = -1

    for diction in dictions:
        count += 1

        axis_options = ['X', 'Y', 'Z']
        for axis in axis_options:
            if axis != symm_axis[0] and max(diction[axis]) != 0:
                radial_axis = axis

        saved_radial_base = []
        saved_radial_top = []
        radial_top = None
        radial_base = None

        for j in range(len(diction[radial_axis[0]]) - 1):
            if diction[radial_axis[0]][j] < 1e-3:
                radial_base = diction[radial_axis[0]][j]
                radial_top = diction[radial_axis[0]][j + 1]
                radial_perc = (radial_loc - diction[radial_axis[0]][j]) / (diction[radial_axis[0]][j + 1] -
                                                                           diction[radial_axis[0]][j])
        for j in range(len(diction[radial_axis[0]]) - 1):
            if str(diction[radial_axis[0]][j])[:6] == str(radial_base)[:6]:
                saved_radial_base.append(j)
            elif str(diction[radial_axis[0]][j])[:6] == str(radial_top)[:6]:
                saved_radial_top.append(j)

        x_axis1 = []
        x_axis2 = []
        x_axis = []

        y_axis1 = []
        y_axis2 = []
        y_axis = []

        for cell in saved_radial_base:
            x_axis1.append(diction[symm_axis[0]][cell])
            y_axis1.append(diction[variable][cell])
        for cell in saved_radial_top:
            x_axis2.append(diction[symm_axis[0]][cell])
            y_axis2.append(diction[variable][cell])
        for coor in range(len(x_axis1)):
            x_axis.append(((x_axis2[coor] - x_axis1[coor]) * radial_perc + x_axis1[coor])/0.0072)
            y_axis.append((y_axis2[coor] - y_axis1[coor]) * radial_perc + y_axis1[coor])

        title_dict = {0: 'DCL'}

        axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                     'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                     'MassfractionofCH4': 'YCH4', 'Axial Velocity[m/s]': 'U(ms-1)'}

        plt.plot(x_axis, y_axis, '.', color=colors[count])
    plt.xlabel('Relative axial location (x/d)')
    if variable == 'Temperature[K]' and variable == 'Axial Velocity[m/s]':
        plt.ylabel('Temperature [K]')
    else:
        plt.ylabel(variable[:4] + ' ' + variable[4:12] + ' ' + variable[12:14] + ' ' + variable[14:] + ' ($Y_{' +
                   variable[14:] + '}$)')
    plt.ylim(ymin=0, ymax=0.175)
    plt.title(radial_axis[0] + '= 0mm (' + title_dict[radial_loc] + ')')

    r_d = ['r/d', 'r/d']
    if variable == 'Axial Velocity[m/s]':
        r_d = ['R/D', 'r/d']
    rpa.plot_desired([title_dict[radial_loc]], r_d, [axis_dict[variable]], category, all_titles,
                     all_data, all_info, rpa.folder_bib_comp, 'radial', path+'/', result_legend)

    plt.gcf().clear()

    return


def integral_plot(path, variable, dictions, axial_locations, symm_axis, result_legend, ref_len):

    all_data, all_info, all_titles, category = rpa.read_pmd_stat()
    if variable == 'Axial Velocity[m/s]':
        all_data, all_info, all_titles, category = rpa.read_tud_ldv()
    count = -1
    colors = ['C0', 'C1', 'C4', 'C5', 'C6', 'C9', 'C7', 'C8']
    EXP_integral = []
    CFD_integral2 = []
    CRN_integral2 = []
    fig, ax1 = plt.subplots()
    ax2 = ax1.twinx()

    for diction in dictions:
        count += 1
        CFD_integral = []
        CRN_integral = []
        maxes = []

        for axial_loc in axial_locations:
            saved_axial_base = []
            saved_axial_top = []
            axial_top = None
            axial_base = None

            for j in range(len(diction[symm_axis[0]])-1):
                if diction[symm_axis[0]][j] <= axial_loc < diction[symm_axis[0]][j+1]:
                    axial_base = diction[symm_axis[0]][j]
                    axial_top = diction[symm_axis[0]][j+1]
                    axial_perc = (axial_loc-diction[symm_axis[0]][j])/(diction[symm_axis[0]][j+1]
                                                                       - diction[symm_axis[0]][j])
            for j in range(len(diction[symm_axis[0]])):
                if str(diction[symm_axis[0]][j])[:6] == str(axial_base)[:6]:
                    saved_axial_base.append(j)
                elif str(diction[symm_axis[0]][j])[:6] == str(axial_top)[:6]:
                    saved_axial_top.append(j)

            x_axis1 = []
            x_axis2 = []
            x_axis = []

            y_axis1 = []
            y_axis2 = []
            y_axis = []

            y_axis1_d = []
            y_axis2_d = []
            y_axis_d = []

            dens1 = []
            dens2 = []
            dens = []

            axis_options = ['X', 'Y', 'Z']
            for axis in axis_options:
                if axis != symm_axis[0] and max(diction[axis]) != 0:
                    ax = axis

            for cell in saved_axial_base:
                x_axis1.append(diction[ax][cell])
                y_axis1.append(diction[variable][cell])
                if variable != 'Temperature[K]' and variable != 'Axial Velocity[m/s]':
                    y_axis1_d.append(diction[variable][cell] * diction['Density[kg/m^3]'][cell] *
                                     diction['Axial Velocity[m/s]'][cell] * diction['Cell Axial Surface[m^2]'][cell])
                else:
                    y_axis1_d.append(diction[variable][cell])

            for cell in saved_axial_top:
                x_axis2.append(diction[ax][cell])
                y_axis2.append(diction[variable][cell])
                if variable != 'Temperature[K]' and variable != 'Axial Velocity[m/s]':
                    y_axis2_d.append(diction[variable][cell] * diction['Density[kg/m^3]'][cell] *
                                     diction['Axial Velocity[m/s]'][cell] * diction['Cell Axial Surface[m^2]'][cell])
                else:
                    y_axis2_d.append(diction[variable][cell])

            for coor in range(len(x_axis1)):
                x_axis.append(((x_axis2[coor] - x_axis1[coor]) * axial_perc + x_axis1[coor])/ref_len)
                y_axis.append(((y_axis2[coor] - y_axis1[coor]) * axial_perc + y_axis1[coor]))
                y_axis_d.append(((y_axis2_d[coor] - y_axis1_d[coor]) * axial_perc + y_axis1_d[coor]))

            maxes.append(max(y_axis))

            axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                         'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                         'MassfractionofCH4': 'YCH4', 'MassflowofCarbon': 'Carbon', 'MassflowofHydrogen': 'Hydrogen',
                         'MassflowofOxygen': 'Oxygen', 'MassflowofNitrogen': 'Nitrogen', 'Axial Velocity[m/s]': 'U(ms-1)'}

            title_dict = {0.0072: 'D01', 0.0144: 'D02', 0.0216: 'D03', 0.108: 'D15', 0.216: 'D30', 0.324: 'D45',
                          0.432: 'D60', 0.54: 'D75'}
            r_d = 'r/d'
            if variable == 'Axial Velocity[m/s]':
                r_d = 'R/D'
            exp_x_axis = [item[category.index(r_d)] for item in all_data[all_titles.index(title_dict[axial_loc])]]
            exp_y_axis = [item[category.index(axis_dict[variable])]
                          for item in all_data[all_titles.index(title_dict[axial_loc])]]

            for cfd_path in os.listdir(os.getcwd()):
                if cfd_path.startswith('SFD_'):
                    break

            result_titles, result_category, result_data = \
                rpa.get_cfd_data('radial', cfd_path, axis_dict[variable])

            cfd_axi = result_data[result_category.index(axis_dict[variable])][
                result_titles[result_category.index(axis_dict[variable])].index(title_dict[axial_loc])]

            cfd_y_axis = []
            cfd_axi_ = []

            for axi in range(len(cfd_axi[0])):
                cfd_axi_.append([cfd_axi[0][axi], cfd_axi[1][axi]])

            cfd_axi_ = sorted(cfd_axi_, key=lambda a_entry: a_entry[0])

            for cfd_cell in cfd_axi_:
                cfd_y_axis.append(cfd_cell[1])

            EXP_integral.append(integral(exp_x_axis, exp_y_axis, max(exp_x_axis)))
            CFD_integral.append(integral(sorted(cfd_axi[0]), cfd_y_axis, min(max(x_axis), max(cfd_axi[0]))))
            CFD_integral2.append(integral(sorted(cfd_axi[0]), cfd_y_axis, max(exp_x_axis)))
            if variable == 'Temperature[K]' and variable == 'Axial Velocity[m/s]':
                CRN_integral.append(sum(y_axis_d))
            else:
                CRN_integral.append(integral(x_axis, y_axis_d, min(max(x_axis), max(cfd_axi[0]))))
            CRN_integral2.append(integral(x_axis, y_axis, max(exp_x_axis)))

        ax1.plot([x*ref_len for x in axial_locations], CRN_integral, '.', color=colors[count],
                 label=result_legend[count])
        ax2.plot([x*ref_len for x in axial_locations], maxes, '_', color=colors[count],
                 label=result_legend[count]+' (max)')

    # plt.plot([x*ref_len for x in axial_locations], CFD_integral, '.', color='C3', label='CFD')
    # plt.plot([x*ref_len for x in axial_locations], EXP_integral, color='C2', label='EXP')

    ax1.set_xlabel('Relative Axial location (x/d)')
    if variable != 'Temperature[K]' and variable != 'Axial Velocity[m/s]':
        ax1.set_ylabel('Mass flow of '+variable[14:])
    else:
        ax1.set_ylabel(variable)
    ax2.set_ylabel('Max '+variable)
    if variable != 'Temperature[K]' and variable != 'Axial Velocity[m/s]':
        plt.title('Mass flow of ' + variable[14:])
    else:
        plt.title('Integral average of ' + variable)
    handle1, label1 = ax1.get_legend_handles_labels()
    handle2, label2 = ax2.get_legend_handles_labels()
    plt.legend(handle1 + handle2, label1 + label2, loc=1, bbox_to_anchor=(1.9, 1), ncol=1)
    if variable == 'MassfractionofNO':
        ax1.set_ylim((0, 0.00005))
    fig.savefig(path + axis_dict[variable] + '_int.png', bbox_inches='tight')

    plt.gcf().clear()

    return [EXP_integral, CFD_integral2, CRN_integral2]


def integral(y, z, max_loc):

    integral = 0
    i = 0
    while y[i] < max_loc and i+1 < len(y):
        if y[i] >= 0:
            if y[i + 1] > max_loc:
                z[i + 1] = (max_loc - y[i]) / (y[i + 1] - y[i]) * (z[i + 1] - z[i])
            try:
                integral += -z[i] * math.pi * y[i] ** 2 + z[i + 1] * math.pi * y[i + 1] ** 2 + math.pi / 3 * (
                            (y[i] ** 3 - y[i + 1] ** 3) * ((z[i + 1] - z[i]) / (y[i + 1] - y[i])))
            except ZeroDivisionError:
                integral += 0

        i += 1

    return integral / (math.pi * max_loc ** 2)


def integral_simple(y, z, max_loc):

    integral = 0
    i = 0
    while y[i] < max_loc and i + 1 < len(y):
        if i == 0:
            integral += 2 * math.pi * y[i] * z[i] * (0.5 * (y[i + 1] + y[i]))
        elif y[i + 1] > max_loc:
            integral += 2 * math.pi * y[i] * z[i] * (0.5 * (max_loc - y[i - 1]))
        else:
            integral += 2 * math.pi * y[i] * z[i] * (0.5 * (y[i + 1] - y[i - 1]))

        i += 1

    return integral / (math.pi * max_loc ** 2)


def cross_contour(path, variable, dictions, axial_locations, symm_axis, result_legend, vmin, vmax, title):
    res_leg = []
    for result in result_legend:
        if result != 'EXP':
            res_leg.append(result.replace(title[4:], '').replace(title[:4], ''))

    for axial_loc in axial_locations:
        values = []
        count = -1

        for diction in dictions:
            count +=1
            saved_axial_base = []
            saved_axial_top = []
            axial_top = None
            axial_base = None

            for j in range(len(diction[symm_axis[0]])-1):
                if diction[symm_axis[0]][j] <= axial_loc < diction[symm_axis[0]][j+1]:
                    axial_base = diction[symm_axis[0]][j]
                    axial_top = diction[symm_axis[0]][j+1]
                    axial_perc = (axial_loc-diction[symm_axis[0]][j])/(diction[symm_axis[0]][j+1]
                                                                       - diction[symm_axis[0]][j])
            for j in range(len(diction[symm_axis[0]])-1):
                if str(diction[symm_axis[0]][j])[:6] == str(axial_base)[:6]:
                    saved_axial_base.append(j)
                elif str(diction[symm_axis[0]][j])[:6] == str(axial_top)[:6]:
                    saved_axial_top.append(j)

            x_axis1 = []
            x_axis2 = []
            x_axis = []

            y_axis1 = []
            y_axis2 = []
            y_axis = []

            axis_options = ['X', 'Y', 'Z']
            for axis in axis_options:
                if axis != symm_axis[0] and max(diction[axis]) != 0:
                    ax = axis

            for cell in saved_axial_base:
                x_axis1.append(diction[ax][cell])
                y_axis1.append(diction[variable][cell])
            for cell in saved_axial_top:
                x_axis2.append(diction[ax][cell])
                y_axis2.append(diction[variable][cell])
            for coor in range(len(x_axis1)):
                x_axis.append(((x_axis2[coor] - x_axis1[coor]) * axial_perc + x_axis1[coor])/0.0072)
                y_axis.append((y_axis2[coor] - y_axis1[coor]) * axial_perc + y_axis1[coor])

            for i in range(20):
                values.append(y_axis)

        axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                     'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                     'MassfractionofCH4': 'YCH4', 'MassflowofCarbon': 'Carbon', 'MassflowofHydrogen': 'Hydrogen',
                     'MassflowofOxygen': 'Oxygen', 'MassflowofNitrogen': 'Nitrogen', 'Axial Velocity[m/s]': 'U(ms-1)',
                     'Density[kg/m^3]': 'Density(kgm^-3)'}

        title_dict = {0.0072: 'D01', 0.0144: 'D02', 0.0216: 'D03', 0.108: 'D15', 0.216: 'D30', 0.324: 'D45',
                      0.432: 'D60', 0.54: 'D75'}

        for cfd_path in os.listdir(os.getcwd()):
            if cfd_path.startswith('SFD_'):
                break

        result_titles, result_category, result_data = \
            rpa.get_cfd_data('radial', cfd_path, axis_dict[variable])

        cfd_axi = result_data[result_category.index(axis_dict[variable])][
            result_titles[result_category.index(axis_dict[variable])].index(title_dict[axial_loc])]

        cfd_y_axis = []
        cfd_axi_ = []

        for axi in range(len(cfd_axi[0])):
            cfd_axi_.append([cfd_axi[0][axi], cfd_axi[1][axi]])

        cfd_axi_ = sorted(cfd_axi_, key=lambda a_entry: a_entry[0])

        for cfd_cell in cfd_axi_:
            cfd_y_axis.append(cfd_cell[1])

        cfd_y = []

        for i in range(len(cfd_y_axis)-1):
            cfd_y.append(0.5*(cfd_y_axis[i]+cfd_y_axis[i+1]))

        doubles = []

        for i in range(len(sorted(cfd_axi[0])) - 1):
            if sorted(cfd_axi[0])[i] == sorted(cfd_axi[0])[i + 1]:
                doubles.append(i)

        for index in sorted(doubles, reverse=True):
            del cfd_y[index]

        length = len(dictions)

        if len(cfd_y) == len(y_axis):
            for i in range(20):
                values.append(cfd_y)

            length += 1
            res_leg.append('CFD')

        zeniths = np.array(x_axis)
        values = np.array(values)
        thetas = np.radians(np.linspace(0, 360, 20*length))

        r, theta = np.meshgrid(zeniths, thetas)
        fig, ax = plt.subplots(subplot_kw=dict(projection='polar'))
        ax.set_theta_zero_location('N', offset=-360/(2*length))
        plt.thetagrids(np.linspace(360/(2*length), 360-360/(2*length), length), ['']*length)
        lab_angles = list(np.radians(np.linspace(180/length, 360-360/(2*length), length)))

        for lab_angle in lab_angles:
            rot_angle = lab_angles.index(lab_angle)*360/length
            if 270 > rot_angle > 90:
                rot_angle = rot_angle+180
            ax.text(lab_angle, 1.1*zeniths.max(), res_leg[lab_angles.index(lab_angle)], va='center', ha='center',
                    rotation=rot_angle, rotation_mode='anchor')

        divides = np.linspace(0, 360-360/length, length)
        for divide in divides:
            ax.plot((0, np.radians(divide)), (0, max(zeniths)), c='w')
        cax = ax.contourf(theta, r, values, np.linspace(vmin, vmax+0.01, 100), cmap='plasma')
        ax.tick_params(axis='y', colors='w')
        cb = fig.colorbar(cax)
        cb.set_label(variable)
        cb.ax.axhline(y=(values.min()-vmin)/(vmax-vmin), color='w')
        cb.ax.axhline(y=(values.max()-vmin)/(vmax-vmin), color='w')

        title_dict = {0.0072: 'D01', 0.0144: 'D02', 0.0216: 'D03', 0.108: 'D15', 0.216: 'D30', 0.324: 'D45',
                      0.432: 'D60', 0.54: 'D75'}

        axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                     'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                     'MassfractionofCH4': 'YCH4', 'MassflowofCarbon': 'Carbon', 'MassflowofHydrogen': 'Hydrogen',
                     'MassflowofOxygen': 'Oxygen', 'MassflowofNitrogen': 'Nitrogen', 'Axial Velocity[m/s]': 'U(ms-1)'}

        ax.set_title(title + ', ' + symm_axis[0] + '=' + str(round(axial_loc * 1000., 1)) + 'mm (' +
                     title_dict[axial_loc] + ')')
        try:
            fig.savefig(path + '/' + axis_dict[variable] + '/' + title_dict[axial_loc] + '_' + axis_dict[variable]
                        + '_contour.png', bbox_inches='tight')
        except:
            fig.savefig(path + '/' + axis_dict[variable] + '_' + title_dict[axial_loc] + '_contour.png',
                        bbox_inches='tight')

        plt.gcf().clear()
        plt.close('all')
    return


# def cross_contour(path, variable, dictions, axial_locations, symm_axis, result_legend, choice, vmin, vmax):
#
#     for axial_loc in axial_locations:
#         diction = dictions[result_legend.index(choice)]
#         saved_axial_base = []
#         saved_axial_top = []
#         axial_top = None
#         axial_base = None
#
#         for j in range(len(diction[symm_axis[0]])-1):
#             if diction[symm_axis[0]][j] <= axial_loc < diction[symm_axis[0]][j+1]:
#                 axial_base = diction[symm_axis[0]][j]
#                 axial_top = diction[symm_axis[0]][j+1]
#                 axial_perc = (axial_loc-diction[symm_axis[0]][j])/(diction[symm_axis[0]][j+1]
#                                                                    - diction[symm_axis[0]][j])
#         for j in range(len(diction[symm_axis[0]])-1):
#             if str(diction[symm_axis[0]][j])[:6] == str(axial_base)[:6]:
#                 saved_axial_base.append(j)
#             elif str(diction[symm_axis[0]][j])[:6] == str(axial_top)[:6]:
#                 saved_axial_top.append(j)
#
#         x_axis1 = []
#         x_axis2 = []
#         x_axis = []
#
#         y_axis1 = []
#         y_axis2 = []
#         y_axis = []
#
#         axis_options = ['X', 'Y', 'Z']
#         for axis in axis_options:
#             if axis != symm_axis[0] and max(diction[axis]) != 0:
#                 ax = axis
#
#         for cell in saved_axial_base:
#             x_axis1.append(diction[ax][cell])
#             y_axis1.append(diction[variable][cell])
#         for cell in saved_axial_top:
#             x_axis2.append(diction[ax][cell])
#             y_axis2.append(diction[variable][cell])
#         for coor in range(len(x_axis1)):
#             x_axis.append(((x_axis2[coor] - x_axis1[coor]) * axial_perc + x_axis1[coor])/0.0072)
#             y_axis.append((y_axis2[coor] - y_axis1[coor]) * axial_perc + y_axis1[coor])
#
#         thetas = np.radians(np.linspace(0, 360, 20))
#         zeniths = np.array(x_axis)
#         values = []
#         for theta in thetas:
#             values.append(y_axis)
#         values = np.array(values)
#
#         r, theta = np.meshgrid(zeniths, thetas)
#         fig, ax = plt.subplots(subplot_kw=dict(projection='polar'))
#         cax = ax.contourf(theta, r, values, np.linspace(vmin, vmax, 100), cmap='plasma')
#         cb = fig.colorbar(cax)
#         cb.set_label(variable)
#         cb.ax.axhline(y=(values.min()-vmin)/(vmax-vmin), color='w')
#         cb.ax.axhline(y=(values.max()-vmin)/(vmax-vmin), color='w')
#
#         title_dict = {0.0072: 'D01', 0.0144: 'D02', 0.0216: 'D03', 0.108: 'D15', 0.216: 'D30', 0.324: 'D45',
#                       0.432: 'D60', 0.54: 'D75'}
#
#         axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
#                      'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
#                      'MassfractionofCH4': 'YCH4', 'MassflowofCarbon': 'Carbon', 'MassflowofHydrogen': 'Hydrogen',
#                      'MassflowofOxygen': 'Oxygen', 'MassflowofNitrogen': 'Nitrogen'}
#
#         ax.set_title(choice + ', ' + symm_axis[0] + '=' + str(round(axial_loc * 1000., 1)) + 'mm (' +
#                      title_dict[axial_loc] + ')')
#         try:
#             fig.savefig(path + '/' + axis_dict[variable] + '/' + title_dict[axial_loc] + '_' + axis_dict[variable]
#                         + '_contour.png', bbox_inches='tight')
#         except:
#             fig.savefig(path + '/' + axis_dict[variable] + '_' + title_dict[axial_loc] + '_contour.png',
#                         bbox_inches='tight')
#
#         plt.gcf().clear()
#         plt.close('all')
#     return

def parallel_contour(path, variable, dictions, result_legend, choice, axial_locations):

    diction = dictions[result_legend.index(choice)]
    x = []
    y = []
    z = []
    start = 0

    for j in range(1, len(diction['X'])):
        if str(diction['X'][j-1])[:6] != str(diction['X'][j])[:6]:
            x.append(diction['X'][start:j])
            y.append(diction['Y'][start:j])
            z.append(diction[variable][start:j])
            start = j

    fig, ax = plt.subplots()
    cax = ax.tricontourf(diction['X'], diction['Y'], diction[variable],
                         np.linspace(diction[variable].min(), diction[variable].max(), 300),
                         cmap='plasma')
    if variable == 'Reactor':
        cax2 = ax.tricontour(diction['X'], diction['Y'], diction[variable], colors='black', linewidths=0.01,
                             levels=np.arange(1.5, int(max(diction['Reactor'])) + 0.5, 1.0))
    ax.set_aspect('equal')
    cb = fig.colorbar(cax, orientation='horizontal')
    cb.ax.set_xticklabels(cb.ax.get_xticklabels(), rotation=45)
    cb.set_label(variable)

    axis_dict = {'Temperature[K]': 'T(K)', 'MassfractionofCO2': 'YCO2', 'MassfractionofH2O': 'YH2O',
                 'MassfractionofCO': 'YCO', 'MassfractionofNO': 'YNO', 'MassfractionofO2': 'YO2',
                 'MassfractionofCH4': 'YCH4', 'MassflowofCarbon': 'Carbon', 'MassflowofHydrogen': 'Hydrogen',
                 'MassflowofOxygen': 'Oxygen', 'MassflowofNitrogen': 'Nitrogen', 'Axial Velocity[m/s]': 'U(ms-1)',
                 'Density[kg/m^3]': 'Density(kgm^-3)', 'Reactor': 'Reactor'}

    ax.set_title(choice)
    for axial_loc in axial_locations:
        ax.axvline(x=axial_loc, linewidth=1, color='w')
    try:
        fig.savefig(path + '/' + axis_dict[variable] + '/' + axis_dict[variable] + '_contour.png', bbox_inches='tight')
    except:
        fig.savefig(path + '/' + axis_dict[variable] + '_contour.png', bbox_inches='tight')

    plt.gcf().clear()
    plt.close('all')

    return diction[variable].min(), diction[variable].max()


def axial_plot0(variable, diction, axial_locations, symm_axis):

    for axial_loc in axial_locations:
        saved_axial_base = []
        saved_axial_top = []
        axial_top = None
        axial_base = None

        for j in range(len(diction[symm_axis[0]])-1):
            if diction[symm_axis[0]][j] <= axial_loc < diction[symm_axis[0]][j+1]:
                axial_base = diction[symm_axis[0]][j]
                axial_top = diction[symm_axis[0]][j+1]
                axial_perc = (axial_loc-diction[symm_axis[0]][j])/(diction[symm_axis[0]][j+1]-diction[symm_axis[0]][j])
        for j in range(len(diction[symm_axis[0]])-1):
            if str(diction[symm_axis[0]][j])[:6] == str(axial_base)[:6]:
                saved_axial_base.append(j)
            elif str(diction[symm_axis[0]][j])[:6] == str(axial_top)[:6]:
                saved_axial_top.append(j)

        x_axis1 = []
        x_axis2 = []
        x_axis = []

        y_axis1 = []
        y_axis2 = []
        y_axis = []

        axis_options = ['X', 'Y', 'Z']
        for axis in axis_options:
            if axis != symm_axis[0] and max(diction[axis]) != 0:
                ax = axis

        for cell in saved_axial_base:
            x_axis1.append(diction[ax][cell])
            y_axis1.append(diction[variable][cell])
        for cell in saved_axial_top:
            x_axis2.append(diction[ax][cell])
            y_axis2.append(diction[variable][cell])
        for coor in range(len(x_axis1)):
            x_axis.append(((x_axis2[coor] - x_axis1[coor]) * axial_perc + x_axis1[coor]))
            y_axis.append((y_axis2[coor] - y_axis1[coor]) * axial_perc + y_axis1[coor])

        plt.plot(x_axis, y_axis)
        plt.xlabel('Radial location (m)')
        plt.ylabel(variable)
        plt.title(symm_axis[0] + '=' + str(round(axial_loc * 1000., 1)) + 'mm')

        plt.gcf().clear()

    return


if __name__ == "__main__":
    loc = os.getcwd()
    path = loc+'/20_19_Static Temperature_X-Coordinate__'
    axial_locations = [0.0072, 0.0144, 0.0216, 0.108, 0.216, 0.324, 0.432, 0.54]
    symm_axis = 'X-Coordinate'
    variable = 'Temperature[K]'

    diction = Post(path, 15, 'X-Coordinate')
    #axial_plot(variable, diction, axial_locations, symm_axis)

